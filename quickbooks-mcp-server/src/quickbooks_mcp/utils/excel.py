"""
Excel report generation for UltraTax integration.
Creates formatted Excel workbooks with tax mapping data.
"""
import logging
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports for tax preparation."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    async def generate_tax_report(
        self,
        reports_data: Dict[str, Any],
        output_path: str,
        start_date: str,
        end_date: str
    ) -> str:
        """
        Generate tax mapping report Excel file.
        
        Args:
            reports_data: Dictionary with profit_loss and/or balance_sheet data
            output_path: Path to save Excel file
            start_date: Report start date
            end_date: Report end date
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating tax report Excel: {output_path}")
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add cover sheet
            self._create_cover_sheet(wb, start_date, end_date)
            
            # Add P&L sheet if available
            if 'profit_loss' in reports_data:
                self._create_pl_sheet(wb, reports_data['profit_loss'])
            
            # Add Balance Sheet if available
            if 'balance_sheet' in reports_data:
                self._create_bs_sheet(wb, reports_data['balance_sheet'])
            
            # Add tax mapping sheet
            self._create_tax_mapping_sheet(wb, reports_data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Tax report saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating tax report: {str(e)}")
            raise
    
    def _create_cover_sheet(self, wb: openpyxl.Workbook, start_date: str, end_date: str):
        """Create cover sheet with report information."""
        ws = wb.create_sheet("Cover")
        
        # Title
        ws['A1'] = "QuickBooks Tax Mapping Report"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Report period
        ws['A3'] = "Report Period:"
        ws['B3'] = f"{start_date} to {end_date}"
        ws['A3'].font = Font(bold=True)
        
        # Generated date
        ws['A4'] = "Generated:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'].font = Font(bold=True)
        
        # Instructions
        ws['A6'] = "Instructions:"
        ws['A6'].font = Font(bold=True, size=12)
        ws['A7'] = "1. Review the Profit & Loss and Balance Sheet tabs"
        ws['A8'] = "2. Use the Tax Mapping tab to map accounts to tax form lines"
        ws['A9'] = "3. Import the mapped data into UltraTax"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_pl_sheet(self, wb: openpyxl.Workbook, pl_data: Dict[str, Any]):
        """Create Profit & Loss sheet."""
        ws = wb.create_sheet("Profit & Loss")
        
        # Headers
        ws['A1'] = "Account"
        ws['B1'] = "Amount"
        ws['C1'] = "Tax Category"
        
        # Style headers
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        row = 2
        for data_row in pl_data.get('rows', []):
            if data_row['type'] == 'section':
                # Section header
                ws[f'A{row}'] = data_row['name']
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # Section rows
                for sub_row in data_row.get('rows', []):
                    if sub_row and sub_row['type'] == 'data':
                        ws[f'A{row}'] = f"  {sub_row['account']}"
                        ws[f'B{row}'] = sub_row['amount']
                        ws[f'B{row}'].number_format = '$#,##0.00'
                        row += 1
            elif data_row['type'] == 'data':
                ws[f'A{row}'] = data_row['account']
                ws[f'B{row}'] = data_row['amount']
                ws[f'B{row}'].number_format = '$#,##0.00'
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        
        # Add borders
        for row_num in range(1, row):
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row_num}'].border = self.border
    
    def _create_bs_sheet(self, wb: openpyxl.Workbook, bs_data: Dict[str, Any]):
        """Create Balance Sheet."""
        ws = wb.create_sheet("Balance Sheet")
        
        # Headers
        ws['A1'] = "Account"
        ws['B1'] = "Amount"
        ws['C1'] = "Tax Category"
        
        # Style headers
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data (similar to P&L)
        row = 2
        for data_row in bs_data.get('rows', []):
            if data_row['type'] == 'section':
                ws[f'A{row}'] = data_row['name']
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for sub_row in data_row.get('rows', []):
                    if sub_row and sub_row['type'] == 'data':
                        ws[f'A{row}'] = f"  {sub_row['account']}"
                        ws[f'B{row}'] = sub_row['amount']
                        ws[f'B{row}'].number_format = '$#,##0.00'
                        row += 1
            elif data_row['type'] == 'data':
                ws[f'A{row}'] = data_row['account']
                ws[f'B{row}'] = data_row['amount']
                ws[f'B{row}'].number_format = '$#,##0.00'
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        
        # Add borders
        for row_num in range(1, row):
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row_num}'].border = self.border
    
    def _create_tax_mapping_sheet(self, wb: openpyxl.Workbook, reports_data: Dict[str, Any]):
        """Create tax mapping sheet for UltraTax."""
        ws = wb.create_sheet("Tax Mapping")
        
        # Headers
        headers = ['Account', 'Amount', 'Tax Form', 'Line Number', 'Description']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add common tax form mappings
        row = 2
        
        # Income accounts
        ws[f'A{row}'] = "Income Accounts"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Sales Revenue"
        ws[f'C{row}'] = "Form 1120"
        ws[f'D{row}'] = "Line 1a"
        ws[f'E{row}'] = "Gross receipts or sales"
        row += 1
        
        # Expense accounts
        ws[f'A{row}'] = "Expense Accounts"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        expense_mappings = [
            ("Advertising", "Form 1120", "Line 14", "Advertising"),
            ("Auto Expense", "Form 1120", "Line 24", "Other deductions"),
            ("Bank Charges", "Form 1120", "Line 24", "Other deductions"),
            ("Insurance", "Form 1120", "Line 15", "Employee benefit programs"),
            ("Office Supplies", "Form 1120", "Line 24", "Other deductions"),
            ("Rent Expense", "Form 1120", "Line 16", "Rent"),
            ("Utilities", "Form 1120", "Line 24", "Other deductions"),
        ]
        
        for account, form, line, desc in expense_mappings:
            ws[f'A{row}'] = account
            ws[f'C{row}'] = form
            ws[f'D{row}'] = line
            ws[f'E{row}'] = desc
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        
        # Add borders
        for row_num in range(1, row):
            for col_num in range(1, 6):
                ws.cell(row=row_num, column=col_num).border = self.border
    
    def generate_audit_trail(
        self,
        transactions: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """
        Generate audit trail Excel file.
        
        Args:
            transactions: List of transactions
            output_path: Path to save Excel file
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating audit trail Excel: {output_path}")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Trail"
            
            # Headers
            headers = ['Date', 'Type', 'Vendor/Customer', 'Account', 'Amount', 'Description', 'Source']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Add transactions
            for row_num, txn in enumerate(transactions, 2):
                ws.cell(row=row_num, column=1).value = txn.get('date', '')
                ws.cell(row=row_num, column=2).value = txn.get('type', '')
                ws.cell(row=row_num, column=3).value = txn.get('vendor_name', '') or txn.get('customer_name', '')
                ws.cell(row=row_num, column=4).value = txn.get('account_name', '')
                ws.cell(row=row_num, column=5).value = txn.get('amount', 0)
                ws.cell(row=row_num, column=5).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=6).value = txn.get('description', '')
                ws.cell(row=row_num, column=7).value = txn.get('source_file', '')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 30
            
            # Add borders
            for row_num in range(1, len(transactions) + 2):
                for col_num in range(1, 8):
                    ws.cell(row=row_num, column=col_num).border = self.border
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Audit trail saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating audit trail: {str(e)}")
            raise
