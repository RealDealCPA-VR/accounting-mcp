"""
Bank Reconciliation Engine
Fuzzy matching algorithm to reconcile QuickBooks transactions with bank statements.
"""
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
from dataclasses import dataclass, field, asdict
from difflib import SequenceMatcher
import re

logger = logging.getLogger(__name__)


@dataclass
class Transaction:
    """Normalized transaction for matching."""
    id: str
    date: datetime
    amount: float
    description: str
    reference: str = ""
    check_number: str = ""
    source: str = ""  # 'qbo' or 'bank'
    raw_data: Dict[str, Any] = field(default_factory=dict)
    
    @classmethod
    def from_qbo(cls, txn: Dict[str, Any]) -> 'Transaction':
        """Create from QuickBooks transaction."""
        date_str = txn.get('date', txn.get('TxnDate', ''))
        if isinstance(date_str, str):
            date = datetime.strptime(date_str[:10], '%Y-%m-%d')
        else:
            date = date_str
            
        return cls(
            id=str(txn.get('id', txn.get('Id', ''))),
            date=date,
            amount=float(txn.get('amount', txn.get('TotalAmt', 0))),
            description=txn.get('vendor', txn.get('memo', txn.get('PrivateNote', ''))),
            reference=txn.get('DocNumber', ''),
            check_number=txn.get('check_number', ''),
            source='qbo',
            raw_data=txn
        )
    
    @classmethod
    def from_bank(cls, txn: Dict[str, Any]) -> 'Transaction':
        """Create from bank statement transaction."""
        date_str = txn.get('date', '')
        if isinstance(date_str, str):
            # Handle various date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y']:
                try:
                    date = datetime.strptime(date_str[:10], fmt)
                    break
                except ValueError:
                    continue
            else:
                date = datetime.now()
        else:
            date = date_str
            
        return cls(
            id=str(txn.get('id', txn.get('reference', ''))),
            date=date,
            amount=float(txn.get('amount', 0)),
            description=txn.get('description', txn.get('memo', '')),
            reference=txn.get('reference', ''),
            check_number=txn.get('check_number', ''),
            source='bank',
            raw_data=txn
        )


@dataclass
class MatchResult:
    """Result of matching two transactions."""
    qbo_transaction: Transaction
    bank_transaction: Transaction
    confidence: float
    match_details: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'qbo_id': self.qbo_transaction.id,
            'bank_id': self.bank_transaction.id,
            'qbo_date': self.qbo_transaction.date.strftime('%Y-%m-%d'),
            'bank_date': self.bank_transaction.date.strftime('%Y-%m-%d'),
            'qbo_amount': self.qbo_transaction.amount,
            'bank_amount': self.bank_transaction.amount,
            'qbo_description': self.qbo_transaction.description,
            'bank_description': self.bank_transaction.description,
            'confidence': round(self.confidence, 3),
            'match_details': self.match_details
        }


@dataclass
class ReconciliationReport:
    """Complete reconciliation report."""
    account_name: str
    statement_date: str
    qbo_beginning_balance: float = 0.0
    qbo_ending_balance: float = 0.0
    bank_ending_balance: float = 0.0
    
    matched_transactions: List[MatchResult] = field(default_factory=list)
    unmatched_qbo: List[Transaction] = field(default_factory=list)
    unmatched_bank: List[Transaction] = field(default_factory=list)
    
    discrepancies: List[Dict[str, Any]] = field(default_factory=list)
    
    @property
    def matched_count(self) -> int:
        return len(self.matched_transactions)
    
    @property
    def matched_amount(self) -> float:
        return sum(m.qbo_transaction.amount for m in self.matched_transactions)
    
    @property
    def unmatched_qbo_count(self) -> int:
        return len(self.unmatched_qbo)
    
    @property
    def unmatched_qbo_amount(self) -> float:
        return sum(t.amount for t in self.unmatched_qbo)
    
    @property
    def unmatched_bank_count(self) -> int:
        return len(self.unmatched_bank)
    
    @property
    def unmatched_bank_amount(self) -> float:
        return sum(t.amount for t in self.unmatched_bank)
    
    @property
    def is_reconciled(self) -> bool:
        """Check if fully reconciled (no unmatched transactions)."""
        return self.unmatched_qbo_count == 0 and self.unmatched_bank_count == 0
    
    @property
    def difference(self) -> float:
        """Calculate the reconciliation difference."""
        return self.bank_ending_balance - self.qbo_ending_balance
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'summary': {
                'account_name': self.account_name,
                'statement_date': self.statement_date,
                'qbo_ending_balance': self.qbo_ending_balance,
                'bank_ending_balance': self.bank_ending_balance,
                'difference': round(self.difference, 2),
                'is_reconciled': self.is_reconciled,
                'matched_count': self.matched_count,
                'matched_amount': round(self.matched_amount, 2),
                'unmatched_qbo_count': self.unmatched_qbo_count,
                'unmatched_qbo_amount': round(self.unmatched_qbo_amount, 2),
                'unmatched_bank_count': self.unmatched_bank_count,
                'unmatched_bank_amount': round(self.unmatched_bank_amount, 2),
            },
            'matched': [m.to_dict() for m in self.matched_transactions],
            'unmatched_qbo': [
                {
                    'id': t.id,
                    'date': t.date.strftime('%Y-%m-%d'),
                    'amount': t.amount,
                    'description': t.description
                }
                for t in self.unmatched_qbo
            ],
            'unmatched_bank': [
                {
                    'id': t.id,
                    'date': t.date.strftime('%Y-%m-%d'),
                    'amount': t.amount,
                    'description': t.description
                }
                for t in self.unmatched_bank
            ],
            'discrepancies': self.discrepancies
        }


class BankReconciler:
    """
    Fuzzy matching engine for bank reconciliation.
    
    Matching Strategy:
    1. Exact match: Same amount, date within tolerance, high description similarity
    2. Amount match: Exact amount, date within tolerance
    3. Fuzzy match: Amount within tolerance, date within tolerance, description match
    4. Check number match: If check numbers present and match
    """
    
    def __init__(
        self,
        date_tolerance_days: int = 3,
        amount_tolerance: float = 0.01,
        description_threshold: float = 0.6,
        confidence_threshold: float = 0.75
    ):
        """
        Initialize reconciler with matching parameters.
        
        Args:
            date_tolerance_days: Days before/after to consider a date match
            amount_tolerance: Dollar amount tolerance for fuzzy amount matching
            description_threshold: Minimum similarity for description matching (0-1)
            confidence_threshold: Minimum confidence to consider a match (0-1)
        """
        self.date_tolerance = timedelta(days=date_tolerance_days)
        self.amount_tolerance = amount_tolerance
        self.description_threshold = description_threshold
        self.confidence_threshold = confidence_threshold
        
        # Common bank description patterns to normalize
        self.noise_patterns = [
            r'\b(pos|ach|wire|xfer|transfer|debit|credit|card|purchase)\b',
            r'\b\d{4,}\b',  # Long numbers (card numbers, references)
            r'[#*]+\d+',    # Masked numbers
            r'\s+',         # Multiple spaces
        ]
    
    def reconcile(
        self,
        qbo_transactions: List[Dict[str, Any]],
        bank_transactions: List[Dict[str, Any]],
        account_name: str,
        statement_date: str,
        bank_ending_balance: float = 0.0,
        qbo_ending_balance: float = 0.0
    ) -> ReconciliationReport:
        """
        Perform bank reconciliation.
        
        Args:
            qbo_transactions: Transactions from QuickBooks
            bank_transactions: Transactions from bank statement
            account_name: Name of account being reconciled
            statement_date: Statement ending date
            bank_ending_balance: Ending balance per bank statement
            qbo_ending_balance: Ending balance per QuickBooks
            
        Returns:
            ReconciliationReport with matched and unmatched transactions
        """
        logger.info(f"Starting reconciliation for {account_name}")
        logger.info(f"QBO transactions: {len(qbo_transactions)}, Bank transactions: {len(bank_transactions)}")
        
        # Convert to normalized Transaction objects
        qbo_txns = [Transaction.from_qbo(t) for t in qbo_transactions]
        bank_txns = [Transaction.from_bank(t) for t in bank_transactions]
        
        # Track which transactions have been matched
        matched_qbo_ids = set()
        matched_bank_ids = set()
        matches = []
        
        # Phase 1: Check number matching (highest confidence)
        matches.extend(self._match_by_check_number(
            qbo_txns, bank_txns, matched_qbo_ids, matched_bank_ids
        ))
        
        # Phase 2: Exact amount + date matching
        matches.extend(self._match_exact(
            qbo_txns, bank_txns, matched_qbo_ids, matched_bank_ids
        ))
        
        # Phase 3: Fuzzy matching for remaining
        matches.extend(self._match_fuzzy(
            qbo_txns, bank_txns, matched_qbo_ids, matched_bank_ids
        ))
        
        # Collect unmatched transactions
        unmatched_qbo = [t for t in qbo_txns if t.id not in matched_qbo_ids]
        unmatched_bank = [t for t in bank_txns if t.id not in matched_bank_ids]
        
        # Build report
        report = ReconciliationReport(
            account_name=account_name,
            statement_date=statement_date,
            qbo_ending_balance=qbo_ending_balance,
            bank_ending_balance=bank_ending_balance,
            matched_transactions=matches,
            unmatched_qbo=unmatched_qbo,
            unmatched_bank=unmatched_bank
        )
        
        # Add discrepancies
        report.discrepancies = self._find_discrepancies(matches, unmatched_qbo, unmatched_bank)
        
        logger.info(f"Reconciliation complete: {report.matched_count} matched, "
                   f"{report.unmatched_qbo_count} unmatched QBO, "
                   f"{report.unmatched_bank_count} unmatched bank")
        
        return report
    
    def _match_by_check_number(
        self,
        qbo_txns: List[Transaction],
        bank_txns: List[Transaction],
        matched_qbo: set,
        matched_bank: set
    ) -> List[MatchResult]:
        """Match transactions by check number."""
        matches = []
        
        for qbo in qbo_txns:
            if qbo.id in matched_qbo or not qbo.check_number:
                continue
                
            for bank in bank_txns:
                if bank.id in matched_bank or not bank.check_number:
                    continue
                
                if qbo.check_number == bank.check_number:
                    # Verify amount matches
                    if abs(qbo.amount - bank.amount) <= self.amount_tolerance:
                        matches.append(MatchResult(
                            qbo_transaction=qbo,
                            bank_transaction=bank,
                            confidence=0.99,
                            match_details={'method': 'check_number', 'check': qbo.check_number}
                        ))
                        matched_qbo.add(qbo.id)
                        matched_bank.add(bank.id)
                        break
        
        return matches
    
    def _match_exact(
        self,
        qbo_txns: List[Transaction],
        bank_txns: List[Transaction],
        matched_qbo: set,
        matched_bank: set
    ) -> List[MatchResult]:
        """Match by exact amount and date within tolerance."""
        matches = []
        
        for qbo in qbo_txns:
            if qbo.id in matched_qbo:
                continue
            
            best_match = None
            best_confidence = 0
            
            for bank in bank_txns:
                if bank.id in matched_bank:
                    continue
                
                # Check exact amount match
                if abs(qbo.amount - bank.amount) > self.amount_tolerance:
                    continue
                
                # Check date within tolerance
                date_diff = abs((qbo.date - bank.date).days)
                if date_diff > self.date_tolerance.days:
                    continue
                
                # Calculate confidence based on date proximity and description similarity
                date_score = 1.0 - (date_diff / (self.date_tolerance.days + 1))
                desc_score = self._description_similarity(qbo.description, bank.description)
                
                # Exact amount match + close date = high confidence
                confidence = 0.7 + (0.15 * date_score) + (0.15 * desc_score)
                
                if confidence > best_confidence:
                    best_confidence = confidence
                    best_match = bank
            
            if best_match and best_confidence >= self.confidence_threshold:
                matches.append(MatchResult(
                    qbo_transaction=qbo,
                    bank_transaction=best_match,
                    confidence=best_confidence,
                    match_details={
                        'method': 'exact_amount',
                        'date_diff_days': abs((qbo.date - best_match.date).days),
                        'description_similarity': self._description_similarity(qbo.description, best_match.description)
                    }
                ))
                matched_qbo.add(qbo.id)
                matched_bank.add(best_match.id)
        
        return matches
    
    def _match_fuzzy(
        self,
        qbo_txns: List[Transaction],
        bank_txns: List[Transaction],
        matched_qbo: set,
        matched_bank: set
    ) -> List[MatchResult]:
        """Fuzzy matching for remaining transactions."""
        matches = []
        
        for qbo in qbo_txns:
            if qbo.id in matched_qbo:
                continue
            
            best_match = None
            best_confidence = 0
            
            for bank in bank_txns:
                if bank.id in matched_bank:
                    continue
                
                confidence, details = self._calculate_match_confidence(qbo, bank)
                
                if confidence > best_confidence:
                    best_confidence = confidence
                    best_match = (bank, details)
            
            if best_match and best_confidence >= self.confidence_threshold:
                bank, details = best_match
                matches.append(MatchResult(
                    qbo_transaction=qbo,
                    bank_transaction=bank,
                    confidence=best_confidence,
                    match_details={'method': 'fuzzy', **details}
                ))
                matched_qbo.add(qbo.id)
                matched_bank.add(bank.id)
        
        return matches
    
    def _calculate_match_confidence(
        self,
        qbo: Transaction,
        bank: Transaction
    ) -> Tuple[float, Dict[str, Any]]:
        """
        Calculate match confidence between two transactions.
        
        Returns:
            Tuple of (confidence score, match details dict)
        """
        details = {}
        
        # Amount scoring (40% weight)
        amount_diff = abs(qbo.amount - bank.amount)
        if amount_diff <= self.amount_tolerance:
            amount_score = 1.0
        elif amount_diff <= 1.0:
            amount_score = 0.8
        elif amount_diff <= 5.0:
            amount_score = 0.5
        else:
            amount_score = max(0, 1 - (amount_diff / max(abs(qbo.amount), 1)))
        details['amount_score'] = round(amount_score, 3)
        details['amount_diff'] = round(amount_diff, 2)
        
        # Date scoring (30% weight)
        date_diff = abs((qbo.date - bank.date).days)
        if date_diff == 0:
            date_score = 1.0
        elif date_diff <= self.date_tolerance.days:
            date_score = 1.0 - (date_diff / (self.date_tolerance.days + 1))
        else:
            date_score = max(0, 0.5 - (date_diff - self.date_tolerance.days) * 0.1)
        details['date_score'] = round(date_score, 3)
        details['date_diff_days'] = date_diff
        
        # Description scoring (30% weight)
        desc_score = self._description_similarity(qbo.description, bank.description)
        details['description_score'] = round(desc_score, 3)
        
        # Weighted confidence
        confidence = (amount_score * 0.4) + (date_score * 0.3) + (desc_score * 0.3)
        
        return confidence, details
    
    def _description_similarity(self, desc1: str, desc2: str) -> float:
        """
        Calculate similarity between two descriptions.
        Normalizes common banking noise before comparison.
        """
        if not desc1 or not desc2:
            return 0.3  # Neutral score for missing descriptions
        
        # Normalize descriptions
        norm1 = self._normalize_description(desc1)
        norm2 = self._normalize_description(desc2)
        
        if not norm1 or not norm2:
            return 0.3
        
        # Use SequenceMatcher for fuzzy comparison
        similarity = SequenceMatcher(None, norm1, norm2).ratio()
        
        # Bonus for exact word matches
        words1 = set(norm1.split())
        words2 = set(norm2.split())
        if words1 and words2:
            word_overlap = len(words1 & words2) / max(len(words1), len(words2))
            similarity = (similarity + word_overlap) / 2
        
        return similarity
    
    def _normalize_description(self, desc: str) -> str:
        """Normalize description for comparison."""
        if not desc:
            return ""
        
        result = desc.lower()
        
        # Remove common noise patterns
        for pattern in self.noise_patterns:
            result = re.sub(pattern, ' ', result, flags=re.IGNORECASE)
        
        # Clean up whitespace
        result = ' '.join(result.split())
        
        return result.strip()
    
    def _find_discrepancies(
        self,
        matches: List[MatchResult],
        unmatched_qbo: List[Transaction],
        unmatched_bank: List[Transaction]
    ) -> List[Dict[str, Any]]:
        """Identify and categorize discrepancies."""
        discrepancies = []
        
        # Low confidence matches
        for match in matches:
            if match.confidence < 0.85:
                discrepancies.append({
                    'type': 'low_confidence_match',
                    'severity': 'warning',
                    'qbo_id': match.qbo_transaction.id,
                    'bank_id': match.bank_transaction.id,
                    'confidence': match.confidence,
                    'message': f"Low confidence match ({match.confidence:.0%}). Please verify."
                })
        
        # Large unmatched QBO transactions
        for txn in unmatched_qbo:
            if abs(txn.amount) > 100:
                discrepancies.append({
                    'type': 'unmatched_qbo',
                    'severity': 'error' if abs(txn.amount) > 500 else 'warning',
                    'transaction_id': txn.id,
                    'date': txn.date.strftime('%Y-%m-%d'),
                    'amount': txn.amount,
                    'description': txn.description,
                    'message': f"QBO transaction ${txn.amount:.2f} not found in bank statement"
                })
        
        # Large unmatched bank transactions
        for txn in unmatched_bank:
            if abs(txn.amount) > 100:
                discrepancies.append({
                    'type': 'unmatched_bank',
                    'severity': 'error' if abs(txn.amount) > 500 else 'warning',
                    'transaction_id': txn.id,
                    'date': txn.date.strftime('%Y-%m-%d'),
                    'amount': txn.amount,
                    'description': txn.description,
                    'message': f"Bank transaction ${txn.amount:.2f} not found in QuickBooks"
                })
        
        return discrepancies


def generate_reconciliation_excel(
    report: ReconciliationReport,
    output_path: str
) -> str:
    """Generate Excel reconciliation report."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = f"Bank Reconciliation: {report.account_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Statement Date: {report.statement_date}"
    
    # Balances
    ws['A4'] = "Balance Summary"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A5'] = "Bank Ending Balance:"
    ws['B5'] = report.bank_ending_balance
    ws['B5'].number_format = '$#,##0.00'
    ws['A6'] = "QBO Ending Balance:"
    ws['B6'] = report.qbo_ending_balance
    ws['B6'].number_format = '$#,##0.00'
    ws['A7'] = "Difference:"
    ws['B7'] = report.difference
    ws['B7'].number_format = '$#,##0.00'
    if report.difference != 0:
        ws['B7'].font = Font(color="FF0000", bold=True)
    
    # Statistics
    ws['A9'] = "Reconciliation Statistics"
    ws['A9'].font = Font(bold=True, size=12)
    ws['A10'] = "Matched Transactions:"
    ws['B10'] = report.matched_count
    ws['A11'] = "Matched Amount:"
    ws['B11'] = report.matched_amount
    ws['B11'].number_format = '$#,##0.00'
    ws['A12'] = "Unmatched QBO:"
    ws['B12'] = report.unmatched_qbo_count
    ws['A13'] = "Unmatched Bank:"
    ws['B13'] = report.unmatched_bank_count
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # Matched transactions sheet
    ws_matched = wb.create_sheet("Matched")
    headers = ['QBO Date', 'Bank Date', 'QBO Amount', 'Bank Amount', 'QBO Description', 'Bank Description', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws_matched.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    for row, match in enumerate(report.matched_transactions, 2):
        ws_matched.cell(row=row, column=1).value = match.qbo_transaction.date.strftime('%Y-%m-%d')
        ws_matched.cell(row=row, column=2).value = match.bank_transaction.date.strftime('%Y-%m-%d')
        ws_matched.cell(row=row, column=3).value = match.qbo_transaction.amount
        ws_matched.cell(row=row, column=3).number_format = '$#,##0.00'
        ws_matched.cell(row=row, column=4).value = match.bank_transaction.amount
        ws_matched.cell(row=row, column=4).number_format = '$#,##0.00'
        ws_matched.cell(row=row, column=5).value = match.qbo_transaction.description[:50]
        ws_matched.cell(row=row, column=6).value = match.bank_transaction.description[:50]
        ws_matched.cell(row=row, column=7).value = f"{match.confidence:.0%}"
    
    # Unmatched QBO sheet
    if report.unmatched_qbo:
        ws_qbo = wb.create_sheet("Unmatched QBO")
        headers = ['Date', 'Amount', 'Description', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws_qbo.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        for row, txn in enumerate(report.unmatched_qbo, 2):
            ws_qbo.cell(row=row, column=1).value = txn.date.strftime('%Y-%m-%d')
            ws_qbo.cell(row=row, column=2).value = txn.amount
            ws_qbo.cell(row=row, column=2).number_format = '$#,##0.00'
            ws_qbo.cell(row=row, column=3).value = txn.description
            ws_qbo.cell(row=row, column=4).value = txn.reference
    
    # Unmatched Bank sheet
    if report.unmatched_bank:
        ws_bank = wb.create_sheet("Unmatched Bank")
        headers = ['Date', 'Amount', 'Description', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws_bank.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        for row, txn in enumerate(report.unmatched_bank, 2):
            ws_bank.cell(row=row, column=1).value = txn.date.strftime('%Y-%m-%d')
            ws_bank.cell(row=row, column=2).value = txn.amount
            ws_bank.cell(row=row, column=2).number_format = '$#,##0.00'
            ws_bank.cell(row=row, column=3).value = txn.description
            ws_bank.cell(row=row, column=4).value = txn.reference
    
    # Discrepancies sheet
    if report.discrepancies:
        ws_disc = wb.create_sheet("Discrepancies")
        headers = ['Type', 'Severity', 'Message', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws_disc.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        for row, disc in enumerate(report.discrepancies, 2):
            ws_disc.cell(row=row, column=1).value = disc['type']
            ws_disc.cell(row=row, column=2).value = disc['severity']
            ws_disc.cell(row=row, column=3).value = disc['message']
            ws_disc.cell(row=row, column=4).value = disc.get('amount', '')
            if disc['severity'] == 'error':
                ws_disc.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
    
    wb.save(output_path)
    return output_path
